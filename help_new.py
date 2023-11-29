from tkinter import *

import os
import re
import threading
from tkinter import filedialog, messagebox
from tkinter.ttk import Progressbar
from pathlib import Path
import webbrowser
import time
import random
import datetime
# import shutil
# from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
import openpyxl
import pyinputplus as pyip
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl.styles import Font

'''

pip install -i https://pypi.tuna.tsinghua.edu.cn/simple -r requirements.txt

Learning material: 
https://blog.csdn.net/qq_43125235/article/details/125601564



pyinputplus:
1, The min, max, greaterThan, and lessThan Keyword Arguments
2, The blank Keyword Argument
3, The limit, timeout, and default Keyword Arguments
4, The allowRegexes and blockRegexes Keyword Arguments
5, 


response = pyip.inputNum(limit=2, default='N/A')
import pyinputplus as pyip
def f():
    try:
        r = pyip.inputNum(limit=2)
    except pyip.RetryLimitException:
        print('l..out')


bill.zou@kingtrans.com.cn

uc9WKN

unbelievably 5.26 solved git access :  git config --global http.proxy 127.0.0.1:10809

pyinstaller ./teyihelper1.2.py --onefile --add-binary "driver\geckodriver.exe;driver" -i i.ico
pyinstaller ./deep_work.py --onefile --add-binary "driver\geckodriver.exe;driver"






'''


'''


Import the openpyxl module.
Call the openpyxl.load_workbook() function.
Get a Workbook object.
Use the active or sheetnames attributes.
Get a Worksheet object.
Use indexing or the cell() sheet method with row and column keyword arguments.
Get a Cell object.
Read the Cell object’s value attribute.
'''







'''
for i in range(1, 8, 2): # Go through every other row:
    print(i, sheet.cell(row=i, column=2).value)

--------------------------------------------------
3 Ways of iterate cells:

for rowOfCellObjects in sheet['A1':'C3']:
    for cellObj in rowOfCellObjects:
        print(cellObj.coordinate, cellObj.value)
    print('--- END OF ROW ---')


for cellObj in list(sheet.columns)[1]:
    print(cellObj.coordinate, cellObj.value)
print('--- END OF COLUMN ---')


for row in range(2, sheet.max_row + 1):
    data = sheet.cell(row=row, column=2).value
    also_data = sheet['B' + str(row)].value
    print(data)
--------------------------------------------------

wb = openpyxl.Workbook() # Create a blank workbook or openpyxl.load_workbook(path)
wb.save('example_copy.xlsx') # Save the workbook.
wb.create_sheet() # Add a new sheet.
sheetx = wb.create_sheet(index=0, title='First Sheet')
def wb['sheetx']
sheet['A1'].value = 'Hello, world!'
sheet['A1'].font = Font(size=24, italic=True, bold=True, name='Calibri')


from openpyxl.utils import get_column_letter, column_index_from_string
sheet.column_dimensions[get_column_letter(2)].width = 20 # The default column width is 8.43 characters(11 point)/c).
sheet.row_dimensions[1].height = 70 # The default row height is 12.75 points.

sheet.merge_cells('A1:D3') # Merge all these cells.
sheet['A1'] = 'Twelve cells merged together.'
sheet.unmerge_cells('A1:D3') # Split these cells up.



sheet.freeze_panes = 'A2' # Freeze the rows above A2.

sheet.freeze_panes = 'A1' or sheet.freeze_panes = None # No frozen panes


import openpyxl, random
wb = openpyxl.Workbook()
sheet = wb.active
n = 20
for i in range(1, n+1):
    sheet['A' + str(i)] = i + random.randint(1, 100)
ref_obj = openpyxl.chart.Reference(sheet, min_col=1, min_row=1, max_col=1, max_row=n)
series_obj = openpyxl.chart.Series(ref_obj, title='First series')
chart_obj = openpyxl.chart.BarChart()
chart_obj.title = 'My Chart'
chart_obj.append(series_obj)
sheet.add_chart(chart_obj, 'C5')
wb.save('sampleChart.xlsx')




------------------------------------------
About regular expression!

The function that creates Regex objects in Python is re.compile(). This function compiles a regular
expression pattern into a RegexObject, which you can then use for various operations like searching
and matching. Raw strings are often used when creating Regex objects to avoid unintended escape
characters. In regular expressions, backslashes are used to escape special characters. Using a raw
string (prefixing the string with 'r') ensures that backslashes are treated as literal characters
and don't require double escaping.The search() method in Python's re module returns a MatchObject
if it finds a match for the regular expression pattern in the string, or it returns None if no
match is found. To get the actual strings that match the pattern from a MatchObject, you can use
the group() or group(0) method. group(0) returns the entire matched string, and you can access
other matched groups using group(1), group(2), and so on, if you have defined capturing groups in
your regular expression.The findall() method returns a list of strings when there are no capturing
groups (parentheses) in the regular expression. It returns a list of tuples of strings when there
are capturing groups. Each tuple contains the matched substrings for each capturing group. The .
character in a regular expression normally matches any character except a newline. If you pass
re.DOTALL as the second argument to re.compile(), it changes the behavior of the . character to
match any character, including newlines. Passing re.VERBOSE as the second argument to re.compile()
allows you to write more readable and organized regular expressions. It ignores whitespace and
allows you to add comments using #. This makes complex regular expressions easier to understand. It
doesn't affect the pattern matching but helps with the regex's readability.


------------------------------------------
About selenium


import time
# 导入selenium包
from selenium import webdriver
from selenium.webdriver.common.by import By
# 打开指定（Firefox）浏览器
browser = webdriver.Firefox()
# 指定加载页面
browser.get("http://www.csdn.net")
# 通过id属性获取搜索输入框
input_text = browser.find_element(By.ID, "toolbar-search-input")
# 向搜索输入框内输入selenium
input_text.send_keys("selenium")
# 设置停留五秒后执行下一步
time.sleep(5)
# 关闭浏览器
browser.quit()


By.ID
By.NAME
By.CLASS_NAME
By.TAG_NAME
By.LINK_TEXT
By.PARTIAL_LINK_TEXT

.send_keys("selenium")
.click()


'''


excelName = 'output_data.xlsx'
excelName2 = 'File.xlsx'

def saveExcel(wb):
    global excelName
    try:
        wb.save(excelName)
    except:
        excelName = f'output_data{round(random.random(), 6)}.xlsx'
        wb.save(excelName)

def saveExcelNew(wb, name):
    global excelName2
    try:
        wb.save(name)
    except:
        excelName2 = f'File{round(random.random(), 6)}.xlsx'
        wb.save(name)   




def get_sheet(path):
    wb = openpyxl.load_workbook(path)
    sheetnames = wb.sheetnames
    print(f"The workbook[{os.path.basename(path)}]'s sheets' names: {sheetnames}.")
    temp_sheetnames = [j+'-'+str(i+1) for i, j in enumerate(sheetnames)]
    sheetnames_prompt = '\n'.join(temp_sheetnames)+'\n' # s1-1\ns2-2\n
    sheet_chosen_index = pypi.inputChoice([str(i+1) for i in range(len(sheetnames))], blank=True, prompt=f"Which sheet do you choose(blank for active sheet): \n{sheetnames_prompt}")
    if sheet_chosen_index:
        sheet_chosen = wb[sheetnames[int(sheet_chosen_index)-1]]
    else:
        sheet_chosen = wb.active
    print(f'The sheet chosen is {sheet_chosen.title}, the highest row number is {sheet_chosen.max_row}, the highest column number is {sheet_chosen.max_column}.')
    return sheet_chosen

def get_cell(c):
    '''Others like: sheet['A1':'C3']; list(sheet.columns)[1] '''
    print('Row %s, Column %s (%s) is %s' % (c.row, c.column,c.coordinate, c.value))




def resource_path(relative_path):
    try:
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.dirname(__file__)
    return os.path.join(base_path, relative_path)


randomTag = ''

outCounter = 2

def process_g():
    ''' main logic'''
    global randomTag
    global stop
    
    # saveExcName = f'tradewolf{now.hour}-{now.minute}.xlsx'
    global saveExcName
    # print('Extracted pages: ', num)
    global num

    driver = webdriver.Firefox()
    driver.implicitly_wait(30)


    driver.get('https://crm2.waimaolang.cn/#/main/find/index/findSocialMedia/Facebook/led/global') # usa
    # driver.find_element(by=By.CSS_SELECTOR, value='input.el-input__inner[placeholder="用户名"]').send_keys('bill')
    driver.find_element(by=By.CSS_SELECTOR, value='input.el-input__inner[placeholder="用户名"]').send_keys('dido')
    driver.find_element(by=By.CSS_SELECTOR, value='input.el-input__inner[placeholder="公司名称"]').send_keys('深圳市经纬集运国际货运代理有限公司')
    # driver.find_element(by=By.CSS_SELECTOR, value='input.el-input__inner[type="password"]').send_keys('kingtrans123') # password  changes
    # driver.find_element(by=By.CSS_SELECTOR, value='input.el-input__inner[type="password"]').send_keys('waimaolang2022')
    driver.find_element(by=By.CSS_SELECTOR, value='input.el-input__inner[type="password"]').send_keys('kingtrans123')
    # driver.find_element(by=By.CSS_SELECTOR, value='button.login-btn').click()

    global signal
    global stop
    stop = False
    
    while not signal:
        time.sleep(0.7)


    try:
        num = int(bill.get())
        print(f'Pages: {num}')
    except:
        print("Type error, pages numbers set to 1")
        num = 1

    defaultPage = num

    startTime = time.time()

    wb = openpyxl.Workbook()


    # defaultPage = 2
    counter = 0

    while True:

        try:
            currentPageDataList = list() # storing the cacheText

            infoCard = driver.find_elements(by=By.CSS_SELECTOR, value='div.public-bottom-list-item')
            print('total cards: ', len(infoCard))
            for i in infoCard: # for each card, make a string represent unit data
                title = i.find_element(by=By.CSS_SELECTOR, value='a').text
                innerInfoList = i.find_elements(by=By.CSS_SELECTOR, value='div.public-bottom-list-item-text>div')
                cacheText ='%'
                cacheText += (title+'%')
                for div in innerInfoList:
                    cacheText += (div.text+'%')
                currentPageDataList.append(cacheText)
            # storeData2Excel(cacheText, wb) 
                print(cacheText)
            storeData2Excel(currentPageDataList, wb)
            try:
                saveExcelNew(wb, saveExcName)
            except:
                if not randomTag:
                    randomTag = str(random.randint(1, 100000))

                saveExcelNew(wb, saveExcName.split('.')[0]+randomTag + '.xlsx')
            # click next page
            # driver.find_element(by=By.CSS_SELECTOR, value='button.btn-next').click()
            driver.execute_script('arguments[0].click()', driver.find_element(by=By.CSS_SELECTOR, value='button.btn-next'))
            time.sleep(random.random()*17)
            time.sleep(2)

            print('CLicked!!')
            counter += 1
            if counter >= defaultPage:
                break
            if stop:
                print("Stopped.")
                break
        except:
            counter += 1
            if counter >= defaultPage:
                break
            if stop:
                print("Stopped.")
                break
                
    print('END')
    f3()
    driver.close()
    saveFile(True)
    f1g()


def storeData2Excel(currentPageDataList, wb):
    titleListA = [
    '标题',
    '邮箱地址',
    '联系电话',
    '公司主页',
    '所属分类',
    '地区/地点',
    '产品介绍',
    '公司简介',
    ]
    titleListB = [
    '标题',
    '邮箱地址',
    '联系电话',
    '公司名称',
    '主要行业',
    '职位/等级',
    '公司网址',
    '个人标签',
    ]
    titleListC = [
    '标题',
    '邮箱地址',
    '联系电话',
    '公司主页',
    '所属分类',
    '地区/地点',
    'ta的全名',
    '个人标签',
    '自我介绍',
    ]
    if '公司简介' in currentPageDataList[0] and '地区/地点' in currentPageDataList[0]:
        print('FACEBOOK')
        print('wait...')
        storeData2ExcelA(currentPageDataList, wb, titleListA)
    elif '个人标签' in currentPageDataList[0] and '职位/等级' in currentPageDataList[0]:
        print('Linkedin')
        print('wait...')
        storeData2ExcelA(currentPageDataList, wb, titleListB)
    elif 'ta的全名' in currentPageDataList[0] and '自我介绍' in currentPageDataList[0]:
        print('INS')
        print('wait...')
        storeData2ExcelA(currentPageDataList, wb, titleListC)



def storeData2ExcelA(currentPageDataList, wb, titleList):
    global outCounter

    sheet = wb.active
    tRe = re.compile(r'[^%]*')

    # first set header

    c = 0
    for t in titleList:
        sheet[chr(ord('A')+c)+str(1)] = t
        sheet[chr(ord('A')+c)+str(1)].font = Font(size=14, bold=True)
        c += 1

    for eachUnit in currentPageDataList:

        # print(txt)
        # storing pcs of into to lt
        lt = []

        for i in tRe.findall(eachUnit):
            # print(i)
            lt.append(i)
            lt = [i for i in lt if i]

        try:
            if len(titleList) == 8:
                dataList = [
                lt[0].strip()
                ,lt[1][4:].strip()
                ,lt[2][4:].strip()
                ,lt[3][4:].strip()
                ,lt[4][4:].strip()
                ,lt[5][5:].strip()
                ,lt[6][4:].strip()
                ,lt[7][4:].strip()

                ]
                c = 0
                for d in dataList:
                    sheet[chr(ord('A')+c)+str(outCounter)] = d
                    c += 1
            elif len(titleList) == 9:
                dataList = [
                lt[0].strip()
                ,lt[1][4:].strip()
                ,lt[2][4:].strip()
                ,lt[3][4:].strip()
                ,lt[4][4:].strip()
                ,lt[5][5:].strip()
                ,lt[6][5:].strip()
                ,lt[7][4:].strip()
                ,lt[8][4:].strip()

                ]
                c = 0
                for d in dataList:
                    sheet[chr(ord('A')+c)+str(outCounter)] = d
                    c += 1              
        except:
            pass
        outCounter += 1




def switchButton(b1):
    # pass
    if b1["state"] == "normal":
        b1["state"] = "disabled"
        # b2["text"] = "enable"
    else:
        b1["state"] = "normal"
        # b2["text"] = "disable"



def f1g():
    # switchButton(b1)
    switchButton(b2)
    switchButton(b1g)
    t = threading.Thread(target=process_g) # , args=(filenames, pb, cwd))
    t.start()



def f2():
    switchButton(b2)
    switchButton(b3)
    t = threading.Thread(target=switch) # , args=(filenames, pb, cwd))
    t.start()   

def f3():
    global outCounter
    outCounter = 2
    switchButton(b3)
    # switchButton(b1)
    switchButton(b1g)
    t = threading.Thread(target=end) # , args=(filenames, pb, cwd))
    # print('??')
    t.start()
    # print('dd')   

def switch():
    # print('sw?')
    global signal
    signal = True

def end():
    # print('????')
    global stop
    global signal   
    stop = True
    print('Stop:', stop)
    signal = False

def f4():
    t = threading.Thread(target=saveFile) # , args=(filenames, pb, cwd))
    t.start()


def saveFile(finishedFlag=None):
    global saveExcName
    if finishedFlag is None:
        fileSave = filedialog.asksaveasfilename(initialfile='result', initialdir = f"{Path().cwd() / 'new_clients_leads_project'}", defaultextension='.xlsx', filetypes = [("Excel files",".xlsx")])
        print(f'File save at {fileSave}')
        saveExcName = fileSave
    if finishedFlag is None:
        label_file_explorer.configure(text="File Saved as: "+os.path.basename(saveExcName))
    else:
        label_file_explorer.configure(text="Finished: "+os.path.basename(saveExcName))




signal = False
stop = False


root = Tk()
# root.config(bg='#345') # config color 
root.title('提取数据助手')
# creating a canvas to insert image
# w, h = 500, 420
# root.geometry('330x150')

def fx():
    t = threading.Thread(target=check)
    t.start()

def check():
    global num
    num = v.get()

num = 3
v = IntVar()
v.set(3)


# print(v)

f = Frame(root)
f.pack()

bill_label = Label(f,text='Pages:  ')
bill_label.grid(row=0,column=0,sticky="W",padx=5,pady=5)
# bill_label.pack(side=TOP)
bill = Entry(f)
bill.insert(END,'1')
bill.grid(row=0,column=1,sticky="E",padx=5,pady=5)
# bill.pack(side=TOP)



 # Button(root, text = "Seller-sprite excel files", command = lambda: select_multi_files('B')).grid(column=2)

b1g = Button(f , text = "Firefox(外贸狼)", command = f1g)
b1g.grid(column=0, row=1+1, sticky='WE', padx=100, pady=5, ipadx=30)
b2 = Button(f , text = "Start", command = f2)
b2.grid(column=0, row=2+1, sticky='WE', padx=100, pady=5, ipadx=30)
b3 = Button(f , text = "End", command = f3)
b3.grid(column=0, row=3+1, sticky='WE', padx=100, pady=5, ipadx=30)

save_file = Button(f,text = "Save Location", command = f4)
save_file.grid(column=0, row=4+1, sticky='WE', padx=100, pady=5, ipadx=30)
label_file_explorer = Label(f)
label_file_explorer.grid(column=0, row=5+1)
b2['state'] = 'disabled'
b3['state'] = 'disabled'


now = datetime.datetime.now()
saveExcName = f'tradewolf {now.hour}-{now.minute}({random.randint(1000,9999)}).xlsx'

# mystd = myStdout()  # 实例化重定向类

root.mainloop() 